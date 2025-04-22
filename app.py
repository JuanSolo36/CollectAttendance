import pandas as pd
from flask import Flask, request, send_file, render_template
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import webbrowser
import threading
import os

app = Flask(__name__)

# Colores
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # Tarde / Exceso
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")    # Almuerzo OK
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")     # Extra / entrada única
yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")   # Entrada especial o almuerzo parcial

def procesar_excel(ruta_archivo):
    if not os.path.exists(ruta_archivo):
        print(f"El archivo {ruta_archivo} no existe.")
        return None

    engine = 'openpyxl' if ruta_archivo.endswith('.xlsx') else 'xlrd'
    
    try:
        df = pd.read_excel(ruta_archivo, engine=engine, header=1)
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
        return None
    
    df = df[['ID', 'Nombre', 'Apellido', 'Tiempo']]
    df['Tiempo'] = pd.to_datetime(df['Tiempo'])
    df['Fecha'] = df['Tiempo'].dt.date
    df['Hora'] = df['Tiempo'].dt.time
    
    output_path = os.path.join(os.getcwd(), 'output_separado_por_dia.xlsx')
    dias_unicos = sorted(df['Fecha'].unique())

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for dia in dias_unicos:
            df_dia = df[df['Fecha'] == dia]
            output_df = pd.DataFrame(columns=[
                'ID', 'Nombre', 'Apellido', 'Fecha', 
                'Hora de Entrada', 'Hora de Almuerzo Salida', 
                'Hora de Almuerzo Entrada', 'Hora de Salida'
            ])
            for id_persona, grupo in df_dia.groupby('ID'):
                nombre = grupo.iloc[0]['Nombre']
                apellido = grupo.iloc[0]['Apellido']
                registros = grupo.sort_values('Tiempo')['Tiempo'].tolist()
                registros_filtrados = [registros[0]]

                for t in registros[1:]:
                    if (t - registros_filtrados[-1]) > timedelta(minutes=3):
                        registros_filtrados.append(t)
                
                entrada = almuerzo_salida = almuerzo_entrada = salida = None

                if len(registros_filtrados) == 1:
                    hora_unica = registros_filtrados[0].time()
                    if hora_unica > datetime.strptime('11:00:00', '%H:%M:%S').time():
                        salida = hora_unica
                    else:
                        entrada = hora_unica
                else:
                    for i, t in enumerate(registros_filtrados):
                        hora = t.time()
                        if i == 0:
                            entrada = hora
                        elif i == len(registros_filtrados) - 1:
                            salida = hora
                        elif almuerzo_salida is None and (t - registros_filtrados[i-1]) > timedelta(minutes=45):
                            almuerzo_salida = hora
                        elif almuerzo_salida and almuerzo_entrada is None:
                            almuerzo_entrada = hora

                output_df.loc[len(output_df)] = [
                    id_persona, nombre, apellido, dia, entrada, 
                    almuerzo_salida, almuerzo_entrada, salida
                ]

            output_df.to_excel(writer, index=False, sheet_name=str(dia))
            worksheet = writer.sheets[str(dia)]

            column_widths = {
                "A": 10, "B": 20, "C": 20, "D": 12, 
                "E": 18, "F": 22, "G": 22, "H": 18,
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            for row in worksheet.iter_rows(min_row=2, max_col=8, max_row=len(output_df)+1):
                try:
                    fecha = row[3].value
                    dia_semana = datetime.strptime(str(fecha), '%Y-%m-%d').weekday()
                    cell_entrada = row[4]
                    cell_almuerzo_salida = row[5]
                    cell_almuerzo_entrada = row[6]
                    cell_salida = row[7]

                    def str_to_time(value):
                        return datetime.strptime(str(value), '%H:%M:%S').time()

                    # ------------------ REGLAS DE ENTRADA ------------------
                    if cell_entrada.value:
                        hora_entrada = str_to_time(cell_entrada.value)
                        hora = hora_entrada.hour
                        minuto = hora_entrada.minute

                        # Si es la única hora registrada (entrada única)
                        if cell_entrada.value and not any([cell_almuerzo_salida.value, cell_almuerzo_entrada.value, cell_salida.value]):
                            if hora_entrada <= datetime.strptime('11:00:00', '%H:%M:%S').time():
                                cell_entrada.fill = blue_fill

                        # Entrada a X:30 o X:15
                        if minuto in [15, 30]:
                            cell_entrada.fill = yellow_fill
                        elif minuto in [16, 31]:
                            cell_entrada.fill = red_fill
                        elif minuto in range(0, 3) or minuto >= 50:
                            pass  # Puntual
                        elif minuto > 2:
                            cell_entrada.fill = red_fill

                    # ------------------ REGLAS DE ALMUERZO ------------------
                    if cell_almuerzo_salida.value and cell_almuerzo_entrada.value:
                        salida_almuerzo = str_to_time(cell_almuerzo_salida.value)
                        entrada_almuerzo = str_to_time(cell_almuerzo_entrada.value)
                        tiempo_almuerzo = (
                            datetime.combine(datetime.min, entrada_almuerzo) - datetime.combine(datetime.min, salida_almuerzo)
                        ).seconds / 60
                        if int(tiempo_almuerzo) > 60:
                            cell_almuerzo_salida.fill = red_fill
                            cell_almuerzo_entrada.fill = red_fill
                        else:
                            cell_almuerzo_salida.fill = green_fill
                            cell_almuerzo_entrada.fill = green_fill

                    # ------------------ SALIDA EN CASOS ESPECIALES ------------------
                    if cell_salida.value and not any([cell_almuerzo_salida.value, cell_almuerzo_entrada.value]):
                        hora_salida = str_to_time(cell_salida.value)

                        if dia_semana < 5 and datetime.strptime('12:00:00', '%H:%M:%S').time() <= hora_salida <= datetime.strptime('13:59:00', '%H:%M:%S').time():
                            cell_salida.fill = yellow_fill
                        elif hora_salida <= datetime.strptime('14:50:00', '%H:%M:%S').time():
                            cell_salida.fill = yellow_fill
                        elif dia_semana == 5:  # Sábado
                            cell_salida.fill = blue_fill
                except Exception as e:
                    print(f"Error al aplicar color en la fila {row[0].row}: {e}")

    return output_path

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        
        file_path = os.path.join("uploads", file.filename)
        file.save(file_path)
        
        output_path = procesar_excel(file_path)
        if output_path:
            return send_file(output_path, as_attachment=True)
        else:
            return 'Error procesando el archivo o archivo no encontrado.'
    
    return render_template('upload.html')

def abrir_navegador():
    webbrowser.open("http://127.0.0.1:5000/")

if __name__ == '__main__':
    if not os.path.exists("uploads"):
        os.makedirs("uploads")
    threading.Timer(1.5, abrir_navegador).start()
    app.run(host='0.0.0.0', port=5000, debug=True)
